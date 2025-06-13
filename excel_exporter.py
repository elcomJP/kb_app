import os
import pandas as pd
from datetime import datetime
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtCore import QDate

# Excelエクスポート用ライブラリ
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    def __init__(self):
        pass
    
    def _format_date(self, date):
        """QDateをYYYY/MM/DD形式の文字列に変換"""
        if isinstance(date, QDate):
            return date.toString('yyyy/MM/dd')
        return str(date)

    def _get_report_title(self, date_str):
        """日付文字列から適切なレポートタイトルを決定する"""
        if "～" in date_str:
            start_date, end_date = date_str.split("～")
            # 空白を削除して比較
            if start_date.strip() != end_date.strip():
                return "売上月計表"
        return "売上日計表"

    def _format_menu_number(self, menu_num):
        """メニュー番号から先頭の0を削除する"""
        try:
            # 文字列に変換してから処理
            menu_str = str(menu_num)
            # 先頭の0を削除
            return str(int(menu_str)) if menu_str.isdigit() else menu_str
        except (ValueError, TypeError):
            return str(menu_num)
    
    def export_to_excel(self, data, file_path, shop_name, title, date_str, parent=None):
        """
        データをExcelファイルにエクスポート
        """
        try:
            # データフレームが空かどうかを確認
            if data is None or len(data) == 0:
                print("エクスポートするデータがありません")
                if parent:
                    QMessageBox.warning(parent, "警告", "エクスポートするデータがありません")
                return False
            
            # titleがQDateオブジェクトの場合は文字列に変換する
            if isinstance(title, QDate):
                title = self._format_date(title)
            
            # データのコピーを作成
            try:
                data_copy = data.copy()
            except:
                # DataFrameでない場合、変換を試みる
                try:
                    data_copy = pd.DataFrame(data)
                except:
                    print("データをDataFrameに変換できません")
                    if parent:
                        QMessageBox.warning(parent, "エラー", "データ形式が不正です")
                    return False
            
            # データ列のマッピング
            column_mapping = {
                'T': '集計Ｇ名称',  # グループ名
                'Q': '商品コード',  # メニュー番号
                'R': '論理口座名称',  # メニュー名
                'J': '枚数',  # 数量
                'L': '金額',  # 金額
                'K': '金額符号',  # 金額符号
                'M': 'カード減算額'  # カード減算額
            }
            
            # 列名のマッピングを確認
            for col_letter, col_name in column_mapping.items():
                if col_name not in data_copy.columns:
                    print(f"警告: 列 '{col_name}' がデータに存在しません")
            
            # 新しいExcelワークブックを作成
            wb = openpyxl.Workbook()
            
            # タイトルを日付範囲に基づいて決定
            pdf_title = self._get_report_title(date_str)
            
            # 集計条件に基づいてデータを分類
            # 【現金売上】K列：金額符号が「0」且つ、M列：カード減算額が「0」のとき
            amount_sign_col = '金額符号' if '金額符号' in data_copy.columns else 'K'
            card_deduction_col = 'カード減算額' if 'カード減算額' in data_copy.columns else 'M'

            normal_data = data_copy[
                (data_copy[amount_sign_col] == '0') & 
                (data_copy[card_deduction_col] == '0')
            ].copy() if amount_sign_col in data_copy.columns and card_deduction_col in data_copy.columns else pd.DataFrame(columns=data_copy.columns)

            # 【赤伝】K列：金額符号が「1」且つ、M列：カード減算額が「0」のとき
            # または、K列：金額符号が「1」且つ、M列：カード減算額が「0」以外のとき
            red_data = data_copy[
                (data_copy[amount_sign_col] == '1')
            ].copy() if amount_sign_col in data_copy.columns else pd.DataFrame(columns=data_copy.columns)

            # 【キャッシュレス決済】K列：金額符号が「0」且つ、M列：カード減算額が「0」以外のとき
            cashless_data = data_copy[
                (data_copy[amount_sign_col] == '0') & 
                (data_copy[card_deduction_col] != '0')
            ].copy() if amount_sign_col in data_copy.columns and card_deduction_col in data_copy.columns else pd.DataFrame(columns=data_copy.columns)
            
            # テーブルヘッダー
            table_header = ['グループ名', 'メニュー番号', 'メニュー名', '数量', '金額']

            # シートの順序を変更するため、最初に総括シートを作成
            wb.active.title = "総括"
            ws_overview = wb["総括"]
            
            # 1. 現金売上シート
            ws_normal = wb.create_sheet(title="現金売上")
            
            # ヘッダー情報を追加
            ws_normal.cell(row=1, column=1, value=f"{pdf_title} ")
            ws_normal.cell(row=2, column=1, value=f"店舗名: {shop_name}")
            ws_normal.cell(row=3, column=1, value=f"集計日: {date_str}")
            
            # タイトルセルのスタイル設定
            title_font = Font(size=14, bold=True)
            ws_normal.cell(row=1, column=1).font = title_font
            
            # カテゴリー名を追加
            ws_normal.cell(row=4, column=1, value="【現金売上】")
            category_font = Font(size=11, bold=True)
            ws_normal.cell(row=4, column=1).font = category_font
            
            # テーブルヘッダーを追加
            for col_idx, header in enumerate(table_header, 1):
                cell = ws_normal.cell(row=6, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # 列幅の設定
                if col_idx == 1:  # グループ名
                    ws_normal.column_dimensions[get_column_letter(col_idx)].width = 20
                elif col_idx == 2:  # メニュー番号
                    ws_normal.column_dimensions[get_column_letter(col_idx)].width = 12
                elif col_idx == 3:  # メニュー名
                    ws_normal.column_dimensions[get_column_letter(col_idx)].width = 40
                else:  # 数量と金額
                    ws_normal.column_dimensions[get_column_letter(col_idx)].width = 15
            
            # ヘッダー行に罫線と背景色を設定
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            border = Border(
                top=Side(style='thin'), 
                bottom=Side(style='thin'), 
                left=Side(style='thin'), 
                right=Side(style='thin')
            )
            
            for col_idx in range(1, len(table_header) + 1):
                ws_normal.cell(row=6, column=col_idx).fill = header_fill
                ws_normal.cell(row=6, column=col_idx).border = border
            
            # 現金売上データを追加
            normal_total = {'数量': 0, '金額': 0}
            row_idx = 7  # ヘッダーの次の行から開始
            
            # データがある場合は追加
            row_idx = self._add_group_data_to_sheet(ws_normal, normal_data, row_idx, normal_total)
            
            # データがない場合
            if row_idx == 7:
                for col_idx in range(1, 6):
                    ws_normal.cell(row=row_idx, column=col_idx).border = border
                ws_normal.cell(row=row_idx, column=1, value="データなし")
                row_idx += 1
            
            # 現金売上の総計を追加
            self._add_total_row(ws_normal, row_idx, "現金売上 計", normal_total, is_main_total=True)
            
            # 2. キャッシュレス決済シート
            ws_cashless = wb.create_sheet(title="キャッシュレス決済")
            
            # ヘッダー情報を追加（同様の形式）
            ws_cashless.cell(row=1, column=1, value=f"{pdf_title} ")
            ws_cashless.cell(row=2, column=1, value=f"店舗名: {shop_name}")
            ws_cashless.cell(row=3, column=1, value=f"集計日: {date_str}")
            ws_cashless.cell(row=1, column=1).font = title_font
            
            # カテゴリー名を追加
            ws_cashless.cell(row=4, column=1, value="【キャッシュレス決済】")
            ws_cashless.cell(row=4, column=1).font = category_font
            
            # テーブルヘッダーを追加（列幅設定を含む）
            for col_idx, header in enumerate(table_header, 1):
                cell = ws_cashless.cell(row=6, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = header_fill
                cell.border = border
                # 列幅の設定（現金売上と同様）
                if col_idx == 1:  # グループ名
                    ws_cashless.column_dimensions[get_column_letter(col_idx)].width = 20
                elif col_idx == 2:  # メニュー番号
                    ws_cashless.column_dimensions[get_column_letter(col_idx)].width = 12
                elif col_idx == 3:  # メニュー名
                    ws_cashless.column_dimensions[get_column_letter(col_idx)].width = 40
                else:  # 数量と金額
                    ws_cashless.column_dimensions[get_column_letter(col_idx)].width = 15
            
            # キャッシュレス決済データを追加
            cashless_total = {'数量': 0, '金額': 0}
            row_idx = 7  # ヘッダーの次の行から開始
            
            # データがある場合は追加
            row_idx = self._add_group_data_to_sheet(ws_cashless, cashless_data, row_idx, cashless_total)
            
            # データがない場合
            if row_idx == 7:
                for col_idx in range(1, 6):
                    ws_cashless.cell(row=row_idx, column=col_idx).border = border
                ws_cashless.cell(row=row_idx, column=1, value="データなし")
                row_idx += 1
            
            # キャッシュレス決済の総計を追加
            self._add_total_row(ws_cashless, row_idx, "キャッシュレス決済 計", cashless_total, is_main_total=True)
            
            # 3. 赤伝シート
            ws_red = wb.create_sheet(title="赤伝")
            
            # ヘッダー情報を追加（同様の形式）
            ws_red.cell(row=1, column=1, value=f"{pdf_title} ")
            ws_red.cell(row=2, column=1, value=f"店舗名: {shop_name}")
            ws_red.cell(row=3, column=1, value=f"集計日: {date_str}")
            ws_red.cell(row=1, column=1).font = title_font
            
            # カテゴリー名を追加
            ws_red.cell(row=4, column=1, value="【赤伝】")
            ws_red.cell(row=4, column=1).font = category_font
            
            # テーブルヘッダーを追加（列幅設定を含む）
            for col_idx, header in enumerate(table_header, 1):
                cell = ws_red.cell(row=6, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = header_fill
                cell.border = border
                # 列幅の設定（現金売上と同様）
                if col_idx == 1:  # グループ名
                    ws_red.column_dimensions[get_column_letter(col_idx)].width = 20
                elif col_idx == 2:  # メニュー番号
                    ws_red.column_dimensions[get_column_letter(col_idx)].width = 12
                elif col_idx == 3:  # メニュー名
                    ws_red.column_dimensions[get_column_letter(col_idx)].width = 40
                else:  # 数量と金額
                    ws_red.column_dimensions[get_column_letter(col_idx)].width = 15
            
            # 赤伝データを追加
            red_total = {'数量': 0, '金額': 0}
            row_idx = 7  # ヘッダーの次の行から開始
            
            # データがある場合は追加
            row_idx = self._add_group_data_to_sheet(ws_red, red_data, row_idx, red_total, is_red_slip=True)
            
            # データがない場合
            if row_idx == 7:
                for col_idx in range(1, 6):
                    ws_red.cell(row=row_idx, column=col_idx).border = border
                ws_red.cell(row=row_idx, column=1, value="データなし")
                row_idx += 1
            
            # 赤伝の総計を追加
            self._add_total_row(ws_red, row_idx, "赤伝 計", red_total, is_main_total=True)
            
            # 総括シートの内容を作成
            # ヘッダー情報を追加
            ws_overview.cell(row=1, column=1, value=f"{pdf_title} ")
            ws_overview.cell(row=2, column=1, value=f"店舗名: {shop_name}")
            ws_overview.cell(row=3, column=1, value=f"集計日: {date_str}")
            ws_overview.cell(row=1, column=1).font = title_font
            
            # カテゴリー名を追加
            ws_overview.cell(row=4, column=1, value="【総括】")
            ws_overview.cell(row=4, column=1).font = category_font
            
            # テーブルヘッダーを追加
            for col_idx, header in enumerate(table_header, 1):
                cell = ws_overview.cell(row=6, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = header_fill
                cell.border = border
                # 列幅の設定
                if col_idx == 1:  # グループ名
                    ws_overview.column_dimensions[get_column_letter(col_idx)].width = 20
                elif col_idx == 2:  # メニュー番号
                    ws_overview.column_dimensions[get_column_letter(col_idx)].width = 12
                elif col_idx == 3:  # メニュー名
                    ws_overview.column_dimensions[get_column_letter(col_idx)].width = 40
                else:  # 数量と金額
                    ws_overview.column_dimensions[get_column_letter(col_idx)].width = 15
            
            # 総括データを追加 - 現金売上、キャッシュレス決済、赤伝の合計を表示
            row_idx = 7
            
            # 現金売上の行
            ws_overview.cell(row=row_idx, column=1, value="現金売上")
            ws_overview.cell(row=row_idx, column=4, value=normal_total['数量'])
            ws_overview.cell(row=row_idx, column=5, value=normal_total['金額'])
            
            # スタイル設定
            for col_idx in range(1, 6):
                cell = ws_overview.cell(row=row_idx, column=col_idx)
                cell.border = border
            
            # 数値列の右揃えと数値フォーマット
            for col_idx in range(4, 6):
                cell = ws_overview.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'
            
            row_idx += 1
            
            # キャッシュレス決済の行
            ws_overview.cell(row=row_idx, column=1, value="キャッシュレス決済")
            ws_overview.cell(row=row_idx, column=4, value=cashless_total['数量'])
            ws_overview.cell(row=row_idx, column=5, value=cashless_total['金額'])
            
            # スタイル設定
            for col_idx in range(1, 6):
                cell = ws_overview.cell(row=row_idx, column=col_idx)
                cell.border = border
            
            # 数値列の右揃えと数値フォーマット
            for col_idx in range(4, 6):
                cell = ws_overview.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'
            
            row_idx += 1
            
            # 赤伝の行
            ws_overview.cell(row=row_idx, column=1, value="赤伝")
            ws_overview.cell(row=row_idx, column=4, value=red_total['数量'])
            ws_overview.cell(row=row_idx, column=5, value=red_total['金額'])
            
            # スタイル設定
            for col_idx in range(1, 6):
                cell = ws_overview.cell(row=row_idx, column=col_idx)
                cell.border = border
            
            # 数値列の右揃えと数値フォーマット
            for col_idx in range(4, 6):
                cell = ws_overview.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'
            
            row_idx += 1
            
            # 総計の行
            total_counts = normal_total['数量'] + cashless_total['数量']
            total_amount = normal_total['金額'] + cashless_total['金額']
            
            ws_overview.cell(row=row_idx, column=1, value="総計(現金･キャッシュレス決済)")
            ws_overview.cell(row=row_idx, column=4, value=total_counts)
            ws_overview.cell(row=row_idx, column=5, value=total_amount)
            
            # 総計行のスタイル
            total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            for col_idx in range(1, 6):
                cell = ws_overview.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.fill = total_fill
                cell.font = Font(bold=True)
            
            # 数値列の右揃え
            for col_idx in range(4, 6):
                cell = ws_overview.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'
            
            # シートの順序を変更
            wb._sheets = [ws_overview, ws_normal, ws_cashless, ws_red]
            
            # ファイルを保存
            wb.save(file_path)
            return True
                
        except Exception as e:
            print(f"Excel出力エラー: {e}")
            import traceback
            traceback.print_exc()
            if parent:
                QMessageBox.critical(parent, "エラー", f"Excel出力に失敗しました:\n{str(e)}")
            return False

    def _add_group_data_to_sheet(self, worksheet, data, start_row, total_accumulator, is_red_slip=False):
        """
        メニュー名ごとにデータを集計してExcelシートに追加する
        is_red_slip: 赤伝の場合はTrue（数量と金額をマイナス表示）
        """
        if data.empty:
            return start_row
        
        # 列名を設定
        group_name_column = '集計Ｇ名称' if '集計Ｇ名称' in data.columns else 'T'  # T列
        menu_number_column = '商品コード' if '商品コード' in data.columns else 'Q'  # Q列
        menu_name_column = '論理口座名称' if '論理口座名称' in data.columns else 'R'  # R列
        quantity_column = '枚数' if '枚数' in data.columns else 'J'  # J列
        amount_column = '金額' if '金額' in data.columns else 'L'  # L列
        
        # メニュー名ごとに集計するためのデータフレームを作成
        menu_summary = {}
        
        # データを走査してメニュー名ごとに集計
        for _, row in data.iterrows():
            group_name = str(row.get(group_name_column, "")).strip()
            if not group_name:  # グループ名が空の場合、デフォルト値を設定
                group_name = "その他"
                
            menu_num = row.get(menu_number_column, "")
            menu_name = row.get(menu_name_column, "")
            
            # メニュー名をキーとして使用
            menu_key = (group_name, menu_num, menu_name)
            
            # 数値を安全に取得
            try:
                quantity = int(row.get(quantity_column, 0))
            except (ValueError, TypeError):
                quantity = 0
                
            try:
                amount = int(row.get(amount_column, 0))
            except (ValueError, TypeError):
                amount = 0
            
            # 赤伝の場合はマイナス値にする
            if is_red_slip:
                quantity = -quantity
                amount = -amount
            
            # 集計辞書に追加または更新
            if menu_key in menu_summary:
                menu_summary[menu_key]['quantity'] += quantity
                menu_summary[menu_key]['amount'] += amount
            else:
                menu_summary[menu_key] = {
                    'quantity': quantity,
                    'amount': amount
                }
        
        # グループごとのデータを整理
        current_group = None
        group_subtotal = {'数量': 0, '金額': 0}
        row_idx = start_row
        
        # 罫線スタイル
        border = Border(
                top=Side(style='thin'), 
                bottom=Side(style='thin'), 
                left=Side(style='thin'), 
                right=Side(style='thin')
            )
        
        # グループ名でソートし、さらに同じグループ内ではメニュー番号で昇順にする
        sorted_keys = sorted(menu_summary.keys(), key=lambda x: (x[0], x[1]))
        
        for key in sorted_keys:
            group_name, menu_num, menu_name = key
            summary = menu_summary[key]
            
            # 新しいグループの開始
            if current_group != group_name:
                # 前のグループの小計を追加
                if current_group is not None:
                    # 小計行を追加
                    self._add_total_row(worksheet, row_idx, f"{current_group} 計", group_subtotal)
                    row_idx += 1
                    # 小計をリセット
                    group_subtotal = {'数量': 0, '金額': 0}
                
                # 新しいグループの設定
                current_group = group_name
                # グループ名を追加
                cell = worksheet.cell(row=row_idx, column=1, value=group_name)
                cell.font = Font(bold=True)
                row_idx += 1
            
            # 商品データを追加
            quantity = summary['quantity']
            amount = summary['amount']
            
            worksheet.cell(row=row_idx, column=1, value="")
            worksheet.cell(row=row_idx, column=2, value=self._format_menu_number(menu_num))
            worksheet.cell(row=row_idx, column=3, value=menu_name)
            worksheet.cell(row=row_idx, column=4, value=quantity)
            worksheet.cell(row=row_idx, column=5, value=amount)
            
            # 数値列の右揃えと数値フォーマット
            for col_idx in range(4, 6):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'
            
            # 各セルに罫線を追加
            for col_idx in range(1, 6):
                worksheet.cell(row=row_idx, column=col_idx).border = border
            
            # 小計と合計を更新
            group_subtotal['数量'] += quantity
            group_subtotal['金額'] += amount
            total_accumulator['数量'] += quantity
            total_accumulator['金額'] += amount
            
            row_idx += 1
        
        # 最後のグループの小計を追加
        if current_group:
            self._add_total_row(worksheet, row_idx, f"{current_group} 計", group_subtotal)
            row_idx += 1
        
        return row_idx
    
    def _add_total_row(self, worksheet, row, label, total, is_main_total=False):
        """
        合計行を追加
        is_main_total: Trueの場合は背景色を変える
        """
        worksheet.cell(row=row, column=1, value=label)
        worksheet.cell(row=row, column=4, value=total['数量'])
        worksheet.cell(row=row, column=5, value=total['金額'])
        
        # 罫線スタイル
        border = Border(
            top=Side(style='thin'), 
            bottom=Side(style='thin'), 
            left=Side(style='thin'), 
            right=Side(style='thin')
        )
        
        # 背景色 - 小計行と総計行の両方に適用
        fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # スタイル設定
        for col_idx in range(1, 6):
            cell = worksheet.cell(row=row, column=col_idx)
            cell.border = border
            cell.font = Font(bold=True)
            cell.fill = fill  # すべての小計と総計に背景色を適用
        
        # 数値列の右揃えと数値フォーマット
        for col_idx in range(4, 6):
            cell = worksheet.cell(row=row, column=col_idx)
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = '#,##0'